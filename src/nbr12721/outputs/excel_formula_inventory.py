"""Inventario de formulas do template ABNT NBR 12721:2006."""

from __future__ import annotations

from openpyxl import load_workbook

from ..settings.config import PLANILHA

__all__ = [
    "carregar_inventario_formulas",
    "celula_tem_formula",
    "inventariar_formulas",
]


def celula_tem_formula(valor) -> bool:
    return isinstance(valor, str) and valor.startswith("=")


def inventariar_formulas(caminho_xlsx: str) -> dict:
    """
    Varre o XLSX e retorna inventario de celulas com formula.

    Retorno:
    {
      "arquivo": "...",
      "total_formulas": 123,
      "abas": {
        "QUADRO I": [{"celula": "C42", "formula": "=SUM(...)"}]
      }
    }
    """
    abas: dict[str, list[dict[str, str]]] = {}
    total = 0

    wb = load_workbook(caminho_xlsx, data_only=False, read_only=True)
    try:
        for ws in wb.worksheets:
            formulas_aba: list[dict[str, str]] = []
            for row in ws.iter_rows():
                for cell in row:
                    valor = cell.value
                    if celula_tem_formula(valor):
                        formulas_aba.append(
                            {"celula": cell.coordinate, "formula": valor}
                        )
            if formulas_aba:
                abas[ws.title] = formulas_aba
                total += len(formulas_aba)
    finally:
        wb.close()

    return {
        "arquivo": caminho_xlsx,
        "total_formulas": total,
        "abas": abas,
    }


def carregar_inventario_formulas(caminho_xlsx: str | None = None) -> dict:
    """Inventaria o template padrao ou o caminho informado."""
    destino = caminho_xlsx or PLANILHA
    return inventariar_formulas(destino)
