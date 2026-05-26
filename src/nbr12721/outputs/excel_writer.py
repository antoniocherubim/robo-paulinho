"""Preenchimento da planilha ABNT NBR 12721:2006 (pandas + openpyxl)."""

from __future__ import annotations

import logging
import time
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .excel_mapping import (
    INFO_PRELIMINARES_CELLS,
    QUADRO1_CONFIG,
    QUADRO2_CONFIG,
    QUADRO3_CELLS,
    QUADRO3_PERCENTUAIS,
    QUADRO4B1_CONFIG,
    QUADRO4B_CONFIG,
    QUADRO5_CELLS,
    QUADRO6_CONFIG,
    QUADRO7_CONFIG,
    QUADRO8_CONFIG,
    obter_valor_path,
    resolver_celula,
)
from .excel_tables import tabela_quadro1, tabela_quadro2, tabela_quadro6, tabela_quadro7, tabela_quadro8

logger = logging.getLogger(__name__)

__all__ = ["preencher_planilha"]


def _numero(valor) -> float:
    try:
        return float(valor) if valor else 0.0
    except (TypeError, ValueError):
        return 0.0


def _escrever_dataframe(
    ws,
    df: pd.DataFrame,
    start_row: int,
    col_map: dict[str, int],
    max_rows: int,
) -> int:
    """Escreve linhas do DataFrame; retorna quantidade escrita."""
    if df.empty:
        return 0
    escritas = 0
    for i, row in df.head(max_rows).iterrows():
        r = start_row + escritas
        for col_name, excel_col in col_map.items():
            if col_name not in row.index:
                continue
            valor = row[col_name]
            if isinstance(valor, float) and col_name != "nome" and col_name != "designacao":
                ws.cell(r, excel_col).value = valor
            else:
                ws.cell(r, excel_col).value = valor if valor != "" else None
        escritas += 1
    return escritas


def _preencher_info_preliminares(wb, dados: dict) -> None:
    ws = wb["INFORMAÇÕES PRELIMINARES"]
    inc = dados.get("incorporador", {})
    resp = dados.get("responsavel", {})
    proj = dados.get("projeto", {})
    pp = proj.get("projetoPadrao", {})

    logger.info(
        "Preenchendo informacoes preliminares | incorporador=%s | edificio=%s | cidadeUf=%s",
        inc.get("nome", ""),
        proj.get("nomeEdificio", ""),
        proj.get("cidadeUf", ""),
    )

    for path, cell in INFO_PRELIMINARES_CELLS.items():
        partes = path.split(".")
        if (
            len(partes) == 3
            and partes[0] == "projeto"
            and partes[1] == "projetoPadrao"
        ):
            resolver_celula(ws, cell, "X" if pp.get(partes[2]) else "")
            continue
        valor = obter_valor_path(dados, path)
        if valor is None:
            resolver_celula(ws, cell, "")
        elif isinstance(valor, bool):
            resolver_celula(ws, cell, "X" if valor else "")
        else:
            resolver_celula(ws, cell, valor if valor != "" else "")


def _preencher_quadro1(wb, dados: dict, df: pd.DataFrame) -> None:
    cfg = QUADRO1_CONFIG
    ws = wb[cfg["sheet"]]
    proj = dados.get("projeto", {})
    local = f"{proj.get('localConstrucao', '')} - {proj.get('cidadeUf', '')}"
    resolver_celula(ws, cfg["local_header"], local)
    n = _escrever_dataframe(ws, df, cfg["start_row"], cfg["columns"], cfg["max_rows"])
    logger.info("Preenchendo QUADRO I com %s pavimento(s)", n)


def _preencher_quadro2(wb, df: pd.DataFrame) -> None:
    cfg = QUADRO2_CONFIG
    ws = wb[cfg["sheet"]]
    n = _escrever_dataframe(ws, df, cfg["start_row"], cfg["columns"], cfg["max_rows"])
    logger.info("Preenchendo QUADRO II com %s unidade(s)", n)


def _preencher_quadro3(wb, dados: dict) -> None:
    ws = wb["QUADRO III"]
    q3 = dados.get("quadro3", {})
    pp3 = q3.get("projetoPadrao", {})
    logger.info(
        "Preenchendo QUADRO III | sindicato=%s | mesCub=%s | valorCub=%s",
        q3.get("sindicato", ""),
        q3.get("mesCub", ""),
        q3.get("valorCub", 0),
    )
    for path, cell in QUADRO3_CELLS.items():
        if path.startswith("projetoPadrao."):
            chave = path.split(".", 1)[1]
            resolver_celula(ws, cell, pp3.get(chave, ""))
        else:
            valor = q3.get(path)
            if path == "valorCub":
                resolver_celula(ws, cell, _numero(valor))
            else:
                resolver_celula(ws, cell, valor or "")

    cub = _numero(q3.get("valorCub"))
    pm = _numero(q3.get("percMateriais"))
    po = _numero(q3.get("percMaoObra"))
    if cub > 0 and pm > 0:
        resolver_celula(ws, "P31", round(cub * pm / 100, 2))
    if cub > 0 and po > 0:
        resolver_celula(ws, "P32", round(cub * po / 100, 2))

    for cell, key in QUADRO3_PERCENTUAIS:
        v = _numero(q3.get(key, 0))
        resolver_celula(ws, cell, v if v > 0 else None)

    pc = _numero(q3.get("percConstrutor", 0))
    pi = _numero(q3.get("percIncorporador", 0))
    resolver_celula(ws, "P61", pc / 100 if pc else None)
    resolver_celula(ws, "P62", pi / 100 if pi else None)


def _preencher_quadro4(wb, df_unidades: pd.DataFrame) -> None:
    if df_unidades.empty:
        return
    cfg4b = QUADRO4B_CONFIG
    sub4b = df_unidades.reindex(columns=list(cfg4b["columns"].keys()), fill_value=0)
    _escrever_dataframe(
        wb[cfg4b["sheet"]],
        sub4b,
        cfg4b["start_row"],
        cfg4b["columns"],
        cfg4b["max_rows"],
    )
    cfg4b1 = QUADRO4B1_CONFIG
    sub4b1 = df_unidades.reindex(columns=list(cfg4b1["columns"].keys()), fill_value=0)
    _escrever_dataframe(
        wb[cfg4b1["sheet"]],
        sub4b1,
        cfg4b1["start_row"],
        cfg4b1["columns"],
        cfg4b1["max_rows"],
    )


def _preencher_quadro5(wb, dados: dict) -> None:
    ws = wb["QUADRO V"]
    q5 = dados.get("quadro5", {})
    logger.info(
        "Preenchendo QUADRO V | tipo=%s | garagens=%s",
        q5.get("tipoEdificacao", ""),
        q5.get("garagens", ""),
    )
    for campo, cell in QUADRO5_CELLS.items():
        valor = q5.get(campo, "")
        if valor:
            resolver_celula(ws, cell, valor)


def _preencher_quadro_tabular(wb, df: pd.DataFrame, cfg: dict[str, Any], rotulo: str) -> None:
    ws = wb[cfg["sheet"]]
    if df.empty:
        logger.info("Preenchendo %s: nenhuma linha (template vazio ignorado)", rotulo)
        return
    n = _escrever_dataframe(ws, df, cfg["start_row"], cfg["columns"], cfg["max_rows"])
    logger.info("Preenchendo %s com %s linha(s)", rotulo, n)


def preencher_planilha(dados, modelo, saida):
    inicio = time.monotonic()
    logger.info("Abrindo planilha modelo: %s", modelo)
    wb = load_workbook(modelo)
    logger.info("Abas carregadas: %s", ", ".join(wb.sheetnames))

    df_quadro1 = tabela_quadro1(dados)
    df_quadro2 = tabela_quadro2(dados)

    _preencher_info_preliminares(wb, dados)
    _preencher_quadro1(wb, dados, df_quadro1)
    _preencher_quadro2(wb, df_quadro2)
    _preencher_quadro3(wb, dados)
    _preencher_quadro4(wb, df_quadro2)
    _preencher_quadro5(wb, dados)
    _preencher_quadro_tabular(wb, tabela_quadro6(dados), QUADRO6_CONFIG, "QUADRO VI")
    _preencher_quadro_tabular(wb, tabela_quadro7(dados), QUADRO7_CONFIG, "QUADRO VII")
    _preencher_quadro_tabular(wb, tabela_quadro8(dados), QUADRO8_CONFIG, "QUADRO VIII")

    logger.info("Salvando planilha preenchida: %s", saida)
    wb.save(saida)
    logger.info("Planilha salva com sucesso em %.2fs", time.monotonic() - inicio)
