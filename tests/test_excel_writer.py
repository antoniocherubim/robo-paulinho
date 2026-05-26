import os
import shutil
import tempfile
import unittest

from nbr12721.settings.config import PLANILHA

try:
    from openpyxl import load_workbook

    from nbr12721.outputs.excel_writer import preencher_planilha
except ImportError:
    load_workbook = None
    preencher_planilha = None


def _dados_fixture_excel() -> dict:
    return {
        "incorporador": {"nome": "ACME LTDA", "cnpj": "12.345.678/0001-90", "endereco": ""},
        "responsavel": {"nome": "MARIA SILVA", "crea": "SP-1/D", "art": "", "endereco": ""},
        "projeto": {
            "nomeEdificio": "RESIDENCIAL TESTE",
            "localConstrucao": "LOTE 22",
            "cidadeUf": "Curitiba-PR",
            "projetoPadrao": {"R": True, "CS": False, "CL": False, "CG": False, "CP": False, "CP1Q": False},
            "qtdUnidades": 320,
            "numPavimentos": 22,
            "vagasUA": 0,
            "vagasAcessorio": 22,
            "vagasComum": 55,
            "areaTerreno": 8958.97,
            "dataAprovacao": "",
            "numAlvara": "2457/2023",
            "padraoAcabamento": "",
        },
        "quadro1": {
            "pavimentos": [
                {
                    "nome": "Pavimentos tipo",
                    "areaPrivCobPadrao": 1012.1,
                    "areaPrivCobDifReal": 0,
                    "areaPrivCobDifEquiv": 0,
                    "areaComumNPCobPadrao": 0,
                    "areaComumNPCobDifReal": 0,
                    "areaComumNPCobDifEquiv": 0,
                    "areaComumPCobPadrao": 0,
                    "areaComumPCobDifReal": 0,
                    "areaComumPCobDifEquiv": 0,
                    "qtdPavimentos": 20,
                }
            ],
        },
        "quadro2": {
            "unidades": [
                {
                    "designacao": "Apto 66m2",
                    "areaPrivCobPadrao": 66.02,
                    "areaPrivCobDifReal": 0,
                    "areaPrivCobDifEquiv": 0,
                    "areaComumNPCobPadrao": 0,
                    "areaComumNPCobDifReal": 0,
                    "areaComumNPCobDifEquiv": 0,
                    "qtdUnidades": 160,
                    "outrasAreasPriv": 0,
                    "areaTerrExcl": 0,
                    "areaTerrComum": 0,
                }
            ],
        },
        "quadro3": {"valorCub": 2500.0, "sindicato": "", "mesCub": "", "projetoPadrao": {}},
        "quadro5": {
            "tipoEdificacao": "RESIDENCIAL VERTICAL",
            "numPavimentos": "22",
            "garagens": "55 vagas comuns; 22 vagas duplas",
            "transicao": "Mezanino",
        },
        "quadro6": {
            "equipamentos": [{"nome": "", "tipo": "", "acabamento": "", "detalhes": ""}],
        },
        "quadro7": {
            "acabamentos": [
                {
                    "dependencia": "",
                    "pisos": "",
                    "paredes": "",
                    "tetos": "",
                    "outros": "",
                }
            ],
        },
        "quadro8": {
            "acabamentos": [
                {
                    "dependencia": "",
                    "pisos": "",
                    "paredes": "",
                    "tetos": "",
                    "outros": "",
                }
            ],
        },
    }


@unittest.skipUnless(
    preencher_planilha and os.path.isfile(PLANILHA),
    "openpyxl/pandas ou planilha modelo indisponivel",
)
class TestExcelWriter(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.modelo = os.path.join(self.tmpdir, "modelo.xlsx")
        self.saida = os.path.join(self.tmpdir, "saida.xlsx")
        shutil.copy(PLANILHA, self.modelo)

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _wb_saida(self):
        preencher_planilha(_dados_fixture_excel(), self.modelo, self.saida)
        return load_workbook(self.saida, data_only=True)

    def test_quadro5_garagens_em_f23(self):
        wb = self._wb_saida()
        ws = wb["QUADRO V"]
        self.assertIn("55 vagas comuns", str(ws["F23"].value or ""))
        self.assertIn("22 vagas duplas", str(ws["F23"].value or ""))
        wb.close()

    def test_quadro5_f21_nao_contem_garagens(self):
        wb = self._wb_saida()
        ws = wb["QUADRO V"]
        f21 = str(ws["F21"].value or "")
        self.assertIn("Mezanino", f21)
        self.assertNotIn("vagas comuns", f21.lower())
        self.assertNotIn("vagas duplas", f21.lower())
        wb.close()

    def test_quadro1_area_e_qtd_linha_17(self):
        wb = self._wb_saida()
        ws = wb["QUADRO I"]
        self.assertAlmostEqual(float(ws["C17"].value or 0), 1012.1, places=1)
        self.assertEqual(int(ws["T17"].value or 0), 20)
        wb.close()

    def test_quadros_6_7_8_template_vazio_nao_preenche_linha_12(self):
        wb = self._wb_saida()
        ws6 = wb["QUADRO VI"]
        ws7 = wb["QUADRO VII"]
        ws8 = wb["QUADRO VIII"]
        for ws, cols in (
            (ws6, ("B12", "D12", "F12", "H12")),
            (ws7, ("B12", "D12", "G12", "J12", "L12")),
            (ws8, ("B12", "D12", "G12", "J12", "L12")),
        ):
            for ref in cols:
                valor = ws[ref].value
                if valor is None or valor == "" or valor == 0:
                    continue
                self.assertEqual(
                    valor,
                    None,
                    msg=f"{ws.title} {ref} deveria estar vazio, obteve {valor!r}",
                )
        wb.close()


if __name__ == "__main__":
    unittest.main()
