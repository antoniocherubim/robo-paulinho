import importlib.util
import os
import shutil
import tempfile
import unittest
from pathlib import Path

from nbr12721.settings.config import PLANILHA

try:
    from openpyxl import load_workbook

    from nbr12721.outputs.excel_audit import _valores_equivalentes, auditar_planilha_preenchida
    from nbr12721.outputs.excel_writer import preencher_planilha

    _spec = importlib.util.spec_from_file_location(
        "test_excel_writer",
        Path(__file__).with_name("test_excel_writer.py"),
    )
    _mod_excel_writer = importlib.util.module_from_spec(_spec)
    _spec.loader.exec_module(_mod_excel_writer)
    _dados_fixture_excel = _mod_excel_writer._dados_fixture_excel
except ImportError:
    load_workbook = None
    auditar_planilha_preenchida = None
    preencher_planilha = None
    _valores_equivalentes = None
    _dados_fixture_excel = None


@unittest.skipUnless(
    auditar_planilha_preenchida and os.path.isfile(PLANILHA),
    "dependencias ou planilha modelo indisponiveis",
)
class TestExcelAudit(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.modelo = os.path.join(self.tmpdir, "modelo.xlsx")
        self.saida = os.path.join(self.tmpdir, "saida.xlsx")
        shutil.copy(PLANILHA, self.modelo)

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_tolerancia_numerica(self):
        self.assertTrue(_valores_equivalentes(1012.1, 1012.1000001))
        self.assertTrue(_valores_equivalentes(0, 0))
        self.assertTrue(_valores_equivalentes("", 0))
        self.assertTrue(_valores_equivalentes(None, 0))
        self.assertFalse(_valores_equivalentes(0, None))
        self.assertFalse(_valores_equivalentes(0, ""))
        self.assertFalse(_valores_equivalentes(1012.1, 0))
        self.assertTrue(_valores_equivalentes(0.0, 0.0))

    def test_auditoria_ok_fixture_coerente(self):
        dados = _dados_fixture_excel()
        preencher_planilha(dados, self.modelo, self.saida)
        resultado = auditar_planilha_preenchida(dados, self.saida)
        self.assertEqual(resultado["erro"], "")
        self.assertTrue(resultado["ok"])
        self.assertEqual(resultado["divergencias"], [])
        self.assertGreater(resultado["celulas_verificadas"], 0)

    def test_auditoria_detecta_celula_alterada(self):
        dados = _dados_fixture_excel()
        preencher_planilha(dados, self.modelo, self.saida)
        wb = load_workbook(self.saida)
        wb["QUADRO V"]["F23"] = "VALOR ERRADO"
        wb.save(self.saida)
        wb.close()

        resultado = auditar_planilha_preenchida(dados, self.saida)
        self.assertEqual(resultado["erro"], "")
        self.assertFalse(resultado["ok"])
        campos = {d["campo"] for d in resultado["divergencias"]}
        celulas = {d["celula"] for d in resultado["divergencias"]}
        self.assertIn("quadro5.garagens", campos)
        self.assertIn("QUADRO V!F23", celulas)

    def test_auditoria_planilha_inexistente(self):
        resultado = auditar_planilha_preenchida({}, "/caminho/inexistente/planilha.xlsx")
        self.assertFalse(resultado["ok"])
        self.assertTrue(resultado["erro"])
        self.assertEqual(resultado["celulas_verificadas"], 0)

    def test_campo_tabular_rastreavel(self):
        dados = _dados_fixture_excel()
        preencher_planilha(dados, self.modelo, self.saida)
        wb = load_workbook(self.saida)
        wb["QUADRO I"]["C17"] = 9999
        wb.save(self.saida)
        wb.close()

        resultado = auditar_planilha_preenchida(dados, self.saida)
        campos = [d["campo"] for d in resultado["divergencias"]]
        self.assertIn("quadro1.pavimentos[0].areaPrivCobPadrao", campos)


if __name__ == "__main__":
    unittest.main()
