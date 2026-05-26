"""Auditoria pos-preenchimento: compara JSON extraido com celulas do XLSX."""

from __future__ import annotations

import logging
import os
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .excel_mapping import (
    INFO_PRELIMINARES_CELLS,
    QUADRO1_CONFIG,
    QUADRO2_CONFIG,
    QUADRO5_CELLS,
    obter_valor_path,
)
from .excel_tables import tabela_quadro1, tabela_quadro2

logger = logging.getLogger(__name__)

__all__ = ["auditar_planilha_preenchida"]

_TOLERANCIA_NUMERICA = 0.001


def _celula_vazia(valor) -> bool:
    if valor is None:
        return True
    if isinstance(valor, str):
        return not valor.strip()
    if isinstance(valor, bool):
        return not valor
    return False


def _valores_equivalentes(esperado, encontrado, tol: float = _TOLERANCIA_NUMERICA) -> bool:
    """
    Compara valores lidos do XLSX com o esperado do JSON.
    Zero no XLSX equivale a vazio somente quando o esperado tambem e vazio.
    Numerico esperado 0 compara como numero, nao como vazio.
    """
    if isinstance(esperado, bool) or isinstance(encontrado, bool):
        return bool(esperado) == bool(encontrado)

    if isinstance(esperado, (int, float)) and not _celula_vazia(esperado):
        if _celula_vazia(encontrado):
            return False
        try:
            return abs(float(esperado) - float(encontrado)) <= tol
        except (TypeError, ValueError):
            return False

    esp_vazio = _celula_vazia(esperado)
    enc_vazio = _celula_vazia(encontrado)

    if esp_vazio and enc_vazio:
        return True
    if esp_vazio:
        if isinstance(encontrado, (int, float)) and float(encontrado) == 0:
            return True
        return enc_vazio
    if enc_vazio:
        return False

    try:
        esp_num = float(esperado)
        enc_num = float(encontrado)
        return abs(esp_num - enc_num) <= tol
    except (TypeError, ValueError):
        pass

    return str(esperado).strip() == str(encontrado).strip()


def _valor_esperado_info(path: str, dados: dict) -> Any:
    partes = path.split(".")
    if len(partes) == 3 and partes[0] == "projeto" and partes[1] == "projetoPadrao":
        pp = dados.get("projeto", {}).get("projetoPadrao", {})
        return "X" if pp.get(partes[2]) else ""
    return obter_valor_path(dados, path)


def _ref_planilha(sheet: str, row: int, col: int) -> str:
    return f"{sheet}!{get_column_letter(col)}{row}"


def _ler_celula(ws, ref: str):
    try:
        return ws[ref].value
    except (KeyError, AttributeError):
        return None


def _registrar_divergencia(
    divergencias: list[dict],
    *,
    celula: str,
    campo: str,
    esperado,
    encontrado,
) -> None:
    divergencias.append(
        {
            "celula": celula,
            "campo": campo,
            "esperado": esperado,
            "encontrado": encontrado,
        }
    )


def _auditar_celulas_fixas(
    wb,
    sheet: str,
    celulas: dict[str, str],
    valor_esperado_fn,
    divergencias: list[dict],
) -> int:
    ws = wb[sheet]
    verificadas = 0
    for campo, ref in celulas.items():
        esperado = valor_esperado_fn(campo)
        encontrado = _ler_celula(ws, ref)
        verificadas += 1
        if not _valores_equivalentes(esperado, encontrado):
            _registrar_divergencia(
                divergencias,
                celula=f"{sheet}!{ref}",
                campo=campo,
                esperado=esperado,
                encontrado=encontrado,
            )
    return verificadas


def _auditar_info_preliminares(wb, dados: dict, divergencias: list[dict]) -> int:
    return _auditar_celulas_fixas(
        wb,
        "INFORMAÇÕES PRELIMINARES",
        INFO_PRELIMINARES_CELLS,
        lambda path: _valor_esperado_info(path, dados),
        divergencias,
    )


def _auditar_quadro5(wb, dados: dict, divergencias: list[dict]) -> int:
    ws = wb["QUADRO V"]
    q5 = dados.get("quadro5", {})
    verificadas = 0
    for campo, ref in QUADRO5_CELLS.items():
        esperado = q5.get(campo, "")
        if not esperado:
            continue
        encontrado = _ler_celula(ws, ref)
        verificadas += 1
        if not _valores_equivalentes(esperado, encontrado):
            _registrar_divergencia(
                divergencias,
                celula=f"QUADRO V!{ref}",
                campo=f"quadro5.{campo}",
                esperado=esperado,
                encontrado=encontrado,
            )
    return verificadas


def _auditar_quadro_tabular(
    wb,
    df,
    cfg: dict[str, Any],
    lista_campo: str,
    divergencias: list[dict],
) -> int:
    if df.empty:
        return 0
    ws = wb[cfg["sheet"]]
    verificadas = 0
    for idx in range(len(df)):
        row_num = cfg["start_row"] + idx
        for col_name, excel_col in cfg["columns"].items():
            esperado = df.iloc[idx][col_name]
            encontrado = ws.cell(row_num, excel_col).value
            verificadas += 1
            campo = f"{lista_campo}[{idx}].{col_name}"
            if not _valores_equivalentes(esperado, encontrado):
                _registrar_divergencia(
                    divergencias,
                    celula=_ref_planilha(cfg["sheet"], row_num, excel_col),
                    campo=campo,
                    esperado=esperado,
                    encontrado=encontrado,
                )
    return verificadas


def _auditar_quadro1(wb, dados: dict, divergencias: list[dict]) -> int:
    cfg = QUADRO1_CONFIG
    ws = wb[cfg["sheet"]]
    verificadas = 0
    proj = dados.get("projeto", {})
    local = f"{proj.get('localConstrucao', '')} - {proj.get('cidadeUf', '')}"
    ref_local = cfg["local_header"]
    encontrado_local = _ler_celula(ws, ref_local)
    verificadas += 1
    if not _valores_equivalentes(local, encontrado_local):
        _registrar_divergencia(
            divergencias,
            celula=f"{cfg['sheet']}!{ref_local}",
            campo="projeto.localConstrucao_cidadeUf",
            esperado=local,
            encontrado=encontrado_local,
        )
    verificadas += _auditar_quadro_tabular(
        wb,
        tabela_quadro1(dados),
        cfg,
        "quadro1.pavimentos",
        divergencias,
    )
    return verificadas


def _auditar_quadro2(wb, dados: dict, divergencias: list[dict]) -> int:
    return _auditar_quadro_tabular(
        wb,
        tabela_quadro2(dados),
        QUADRO2_CONFIG,
        "quadro2.unidades",
        divergencias,
    )


def _resultado_base(caminho_xlsx: str, erro: str = "") -> dict:
    return {
        "ok": not erro,
        "divergencias": [],
        "celulas_verificadas": 0,
        "planilha": caminho_xlsx,
        "erro": erro,
    }


def auditar_planilha_preenchida(dados: dict, caminho_xlsx: str) -> dict:
    """
    Compara dados extraidos com celulas preenchidas no XLSX.
    Nao altera arquivos e nao levanta excecao para falhas de leitura.
    """
    if not caminho_xlsx or not os.path.isfile(caminho_xlsx):
        msg = f"planilha nao encontrada: {caminho_xlsx}"
        logger.warning("Auditoria planilha: %s", msg)
        return _resultado_base(caminho_xlsx, erro=msg)

    try:
        wb = load_workbook(caminho_xlsx, data_only=True)
    except Exception as exc:
        msg = f"falha ao abrir planilha: {exc}"
        logger.warning("Auditoria planilha: %s", msg)
        return _resultado_base(caminho_xlsx, erro=msg)

    divergencias: list[dict] = []
    verificadas = 0
    try:
        verificadas += _auditar_info_preliminares(wb, dados, divergencias)
        verificadas += _auditar_quadro5(wb, dados, divergencias)
        verificadas += _auditar_quadro1(wb, dados, divergencias)
        verificadas += _auditar_quadro2(wb, dados, divergencias)
    except Exception as exc:
        wb.close()
        msg = f"falha durante auditoria: {exc}"
        logger.warning("Auditoria planilha: %s", msg)
        return _resultado_base(caminho_xlsx, erro=msg)

    wb.close()
    return {
        "ok": len(divergencias) == 0,
        "divergencias": divergencias,
        "celulas_verificadas": verificadas,
        "planilha": caminho_xlsx,
        "erro": "",
    }
