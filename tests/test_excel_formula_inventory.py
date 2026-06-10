import os
import shutil
import tempfile
import unittest

from nbr12721.settings.config import PLANILHA

try:
    from openpyxl import load_workbook

    from nbr12721.outputs.excel_formula_inventory import (
        carregar_inventario_formulas,
        celula_tem_formula,
        inventariar_formulas,
    )
    from nbr12721.outputs.excel_writer import preencher_planilha
except ImportError:
    load_workbook = None
    carregar_inventario_formulas = None
    celula_tem_formula = None
    inventariar_formulas = None
    preencher_planilha = None


def _dados_fixture_excel() -> dict:
    return {
        "incorporador": {"nome": "ACME LTDA", "cnpj": "12.345.678/0001-90", "endereco": ""},
        "responsavel": {"nome": "MARIA SILVA", "crea": "SP-1/D", "art": "", "endereco": ""},
        "projeto": {
            "nomeEdificio": "RESIDENCIAL TESTE",
            "localConstrucao": "LOTE 22",
            "cidadeUf": "Curitiba-PR",
            "projetoPadrao": {"R": True, "CS": False, "CL": False, "CG": False, "CP": False, "CP1Q": False},
            "qtdUnidades": 320,
            "numPavimentos": 22,
            "vagasUA": 0,
            "vagasAcessorio": 22,
            "vagasComum": 55,
            "areaTerreno": 8958.97,
            "dataAprovacao": "",
            "numAlvara": "2457/2023",
            "padraoAcabamento": "",
        },
        "quadro1": {
            "pavimentos": [
                {
                    "nome": "Pavimentos tipo",
                    "areaPrivCobPadrao": 1012.1,
                    "areaPrivCobDifReal": 0,
                    "areaPrivCobDifEquiv": 0,
                    "areaComumNPCobPadrao": 0,
                    "areaComumNPCobDifReal": 0,
                    "areaComumNPCobDifEquiv": 0,
                    "areaComumPCobPadrao": 0,
                    "areaComumPCobDifReal": 0,
                    "areaComumPCobDifEquiv": 0,
                    "qtdPavimentos": 20,
                }
            ],
        },
        "quadro2": {
            "unidades": [
                {
                    "designacao": "Apto 66m2",
                    "areaPrivCobPadrao": 66.02,
                    "areaPrivCobDifReal": 0,
                    "areaPrivCobDifEquiv": 0,
                    "areaComumNPCobPadrao": 0,
                    "areaComumNPCobDifReal": 0,
                    "areaComumNPCobDifEquiv": 0,
                    "qtdUnidades": 160,
                    "outrasAreasPriv": 0,
                    "areaTerrExcl": 0,
                    "areaTerrComum": 0,
                }
            ],
        },
        "quadro3": {"valorCub": 2500.0, "sindicato": "", "mesCub": "", "projetoPadrao": {}},
        "quadro5": {
            "tipoEdificacao": "RESIDENCIAL VERTICAL",
            "numPavimentos": "22",
            "garagens": "55 vagas comuns; 22 vagas duplas",
            "transicao": "Mezanino",
        },
        "quadro6": {
            "equipamentos": [{"nome": "", "tipo": "", "acabamento": "", "detalhes": ""}],
        },
        "quadro7": {
            "acabamentos": [
                {"dependencia": "", "pisos": "", "paredes": "", "tetos": "", "outros": ""}
            ],
        },
        "quadro8": {
            "acabamentos": [
                {"dependencia": "", "pisos": "", "paredes": "", "tetos": "", "outros": ""}
            ],
        },
    }


class TestCelulaTemFormula(unittest.TestCase):
    @unittest.skipUnless(celula_tem_formula is not None, "openpyxl indisponivel")
    def test_celula_tem_formula(self):
        self.assertTrue(celula_tem_formula("=SUM(A1:A2)"))
        self.assertFalse(celula_tem_formula("texto"))
        self.assertFalse(celula_tem_formula(123))


@unittest.skipUnless(
    inventariar_formulas and os.path.isfile(PLANILHA),
    "openpyxl ou planilha modelo indisponivel",
)
class TestInventariarFormulasTemplate(unittest.TestCase):
    def test_inventariar_formulas_template(self):
        inv = inventariar_formulas(PLANILHA)
        self.assertGreater(inv["total_formulas"], 0)
        for aba in ("QUADRO I", "QUADRO II", "QUADRO III"):
            self.assertIn(aba, inv["abas"])

    def test_inventario_formulas_estrutura(self):
        inv = carregar_inventario_formulas(PLANILHA)
        self.assertIn("arquivo", inv)
        self.assertIn("total_formulas", inv)
        self.assertIn("abas", inv)
        self.assertGreater(inv["total_formulas"], 0)
        primeiro = next(iter(inv["abas"].values()))
        self.assertIn("celula", primeiro[0])
        self.assertIn("formula", primeiro[0])


@unittest.skipUnless(
    preencher_planilha and os.path.isfile(PLANILHA),
    "openpyxl/pandas ou planilha modelo indisponivel",
)
class TestWriterProtegeFormulaTabular(unittest.TestCase):
    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self.modelo = os.path.join(self.tmpdir, "modelo.xlsx")
        self.saida = os.path.join(self.tmpdir, "saida.xlsx")
        shutil.copy(PLANILHA, self.modelo)
        wb = load_workbook(self.modelo)
        wb["QUADRO I"]["C17"] = "=1+1"
        wb.save(self.modelo)
        wb.close()

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_writer_nao_sobrescreve_formula_tabular(self):
        preencher_planilha(_dados_fixture_excel(), self.modelo, self.saida)
        wb = load_workbook(self.saida, data_only=False)
        self.assertEqual(wb["QUADRO I"]["C17"].value, "=1+1")
        wb.close()


if __name__ == "__main__":
    unittest.main()
